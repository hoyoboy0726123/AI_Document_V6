"""Build an Excel (.xlsx) of the 42-question Q&A test results for collection.
Reads qa_results.jsonl + an inline VERDICTS map; writes qa_collection.xlsx."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "qa_results.jsonl")
OUT = r"C:\Users\G635LXG\Downloads\RAG\MIL-STD-810H_QA_collection.xlsx"

# idx -> verdict（過關 / 部分 / 失敗）。未列者預設「過關」。
VERDICTS = {}
PASS_ONLY = False  # True = 只輸出過關題

records = []
with open(SRC, encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if line:
            records.append(json.loads(line))

wb = Workbook()
ws = wb.active
ws.title = "MIL-STD-810H QA"

headers = ["輪次", "編號", "類型", "問題", "路由模式", "工具數", "耗時(秒)", "完整答案", "結果"]
ws.append(headers)
hfill = PatternFill("solid", fgColor="1F4E78")
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = hfill
    cell.alignment = Alignment(horizontal="center", vertical="center")

pass_fill = PatternFill("solid", fgColor="E2EFDA")
part_fill = PatternFill("solid", fgColor="FFF2CC")
fail_fill = PatternFill("solid", fgColor="FBE4D5")
r = 2
n_out = 0
for rec in records:
    verdict = VERDICTS.get(rec["idx"], "過關")
    if PASS_ONLY and verdict != "過關":
        continue
    ws.append([
        f"第{rec['round']}輪", rec["idx"], rec["category"], rec["question"],
        rec["route"], rec["tools"], rec["seconds"], rec["answer"], verdict,
    ])
    fill = {"過關": pass_fill, "部分": part_fill, "失敗": fail_fill}.get(verdict)
    if fill:
        ws.cell(row=r, column=9).fill = fill
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=(c in (4, 8)))
    r += 1
    n_out += 1

widths = [7, 6, 8, 42, 10, 7, 8, 100, 8]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:I{r-1}"
wb.save(OUT)
print(f"wrote {n_out} rows -> {OUT}")
